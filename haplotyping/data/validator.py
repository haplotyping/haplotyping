import os, logging, tempfile
import gzip,csv
from openpyxl import load_workbook
from frictionless import Resource, Schema, Field, Package, Layout
from frictionless.plugins.excel import ExcelDialect
from frictionless import validate, extract
import pandas as pd, numpy as np, textwrap
from shutil import copyfile

class ValidateReport:
    
    def __init__(self, logger:logging.Logger):
        self._logger = logger
        self._reportNames = []
        self._reports = {}
        
    def addReport(self, name:str, forceReport:bool):
        """
        Add report
        """        
        if name in self._reportNames or name in self._reports.keys():
            if forceReport:
                if name in self._reportNames:
                    self._reportNames.remove(name)
                if name in self._reports.keys():
                    del self._reports[name]
            else:
                raise KeyError
        self._reportNames.append(name)
        self._reports[name] = {"valid": True, "errors": [], "warnings": [], "debugs": [], "infos": []}
            
    def addReportError(self, name:str, error:str):
        """
        Add error to report
        """
        if not (name in self._reportNames or name in self._reports.keys()):
            raise KeyError
        else:
            self._reports[name]["valid"] = False
            self._reports[name]["errors"].append(error)
            self._logger.error("[{}] - {}".format(name,error))
            
    def addReportWarning(self, name:str, warning:str):
        """
        Add warning to report
        """
        if not (name in self._reportNames or name in self._reports.keys()):
            raise KeyError
        else:
            self._reports[name]["warnings"].append(warning)
            self._logger.warning("[{}] - {}".format(name,warning))
            
    def addReportDebug(self, name:str, debug:str):
        """
        Add debug to report
        """
        if not (name in self._reportNames or name in self._reports.keys()):
            raise KeyError
        else:
            self._reports[name]["debugs"].append(debug)
            self._logger.debug("[{}] - {}".format(name,debug))
            
    def addReportInfo(self, name:str, info:str):
        """
        Add info to report
        """
        if not (name in self._reportNames or name in self._reports.keys()):
            raise KeyError
        else:
            self._reports[name]["infos"].append(info)
            self._logger.info("[{}] - {}".format(name,info))
            
    def setReportFrictionless(self, name:str, frictionlessReport):
        """
        Set frictionless for report
        """
        if not (name in self._reportNames or name in self._reports.keys()):
            raise KeyError
        else:
            self._reports[name]["frictionless"] = frictionlessReport
        
    @property
    def valid(self):        
        for report in self._reportNames:
            if report in self._reports.keys():
                #check report
                if not self._reports[report]["valid"]:
                    return False
                #check also for frictionless
                if "frictionless" in self._reports[report].keys():
                    if not self._reports[report]["frictionless"].valid:
                        return False                
        return True
    
    @property
    def reports(self):
        return self._reportNames
    
    def __getitem__(self, key):
        if key not in self._reports.keys():
            raise KeyError
        return self._reports[key]
        
        
class ValidateGeneral:

    def __init__(self, dataPath:str, resourceFilename:str, **args):
        
        #get config
        self._resourceFilename = resourceFilename
        self._schemaPath = os.path.join(os.path.dirname(os.path.realpath(__file__)),"schema")
        self._dataPath = dataPath
        self._removeEmptyRows = args.get("removeEmptyRows", True)
        self._removeEmptyColumns = args.get("removeEmptyColumns", True)
        self._adjustTypeForStringColumns = args.get("adjustTypeForStringColumns", True)
        self._allowAdditionalSheets = args.get("allowAdditionalSheets", False)
        
        #logging
        self._logger = logging.getLogger(__name__)
        
        #initialise
        self._report = ValidateReport(self._logger)
        self._package = Package()
        
        self._name = args.get("name", resourceFilename)
        
        self._availableSheets = set()
        self._allowedSheets = set()

            
    def _checkMissingColumns(self, sheetName: str, resourceName:str, skipRows=0):
        if not resourceName in self._report.reports:
            pass
        elif not sheetName in self._availableSheets:
            pass
        elif not self._package.has_resource(resourceName):
            pass
        else:
            try:
                resourceColumnNames = self._getSheetColumnNames(sheetName, skipRows)
                requiredNames = [item["name"] for item in self._package.get_resource(resourceName).schema.fields]
                missingNames = set(requiredNames).difference(resourceColumnNames)
                additionalNames = set(resourceColumnNames).difference(requiredNames)
                if len(missingNames)>0:
                    self._report.addReportError(resourceName,"missing column(s): {}".format(str(list(missingNames))))
                if len(additionalNames)>0:
                    self._report.addReportError(resourceName,"additional column(s): {}".format(str(list(additionalNames))))
                recognizedResourceNames = [item for item in resourceColumnNames if item in requiredNames]
                recognizedRequiredNames = [item for item in requiredNames if item in resourceColumnNames]
                if not (recognizedResourceNames==recognizedRequiredNames):
                    self._report.addReportError(resourceName,"incorrect order column(s), expected: {}".
                                                format(str(requiredNames)))
                    
            except:
                pass
            
    def _ignoreResource(self, sheetName:str):
        self._allowedSheets.add(sheetName)

    def _createEmptyResource(self,sheetName:str, schema, resourceName:str, layout=None):
        #define resource
        self._allowedSheets.add(sheetName)
        self._report.addReport(resourceName, True)
        #try to get schema
        if isinstance(schema, Schema):
            self._report.addReportDebug(resourceName,"use provided schema") 
            resourceSchema = schema
        elif isinstance(schema, str):
            schemaFilename = os.path.join(self._schemaPath,schema)
            if os.access(schemaFilename, os.R_OK):
                self._report.addReportDebug(resourceName,"validate using schema '{}'".format(schema)) 
                resourceSchema = Schema(schemaFilename)
            else:
                self._report.addReportError(resourceName,"could not read schema '{}'".format(schemaFilename))
        else:
            self._report.addReportError(resourceName,"no recognized schema format")
        resource = Resource([[field.name for field in resourceSchema.fields]])
        resource.schema = resourceSchema
        if layout:
            resource.layout = layout
        resource.name = resourceName
        validation_report = validate(resource)
        if not validation_report.valid:
            self._logger.error("invalid {} resource".format(resourceName))   
        #try to add resource
        self._package.add_resource(resource) 
    
    def _validateResource(self, sheetName:str, schema, resourceName:str, headerRows:int=1, skip_errors:list=None):        
        #define resource
        self._allowedSheets.add(sheetName)
        self._report.addReport(resourceName, True)
        if not sheetName in self._availableSheets:
            self._report.addReportError(resourceName,"sheet '{}' not available".format(sheetName))
        else:
            self._report.addReportDebug(resourceName,"define resource '{}'".format(sheetName))
            resource = Resource(basepath=self._dataPath, path=self._resourceFilename, hashing="", 
                                                dialect=ExcelDialect(sheet=sheetName, preserve_formatting=False))  
            if headerRows>1:
                resource.layout=Layout(header_rows=[headerRows])
            #set name and remove if exists
            if self._package.has_resource(resourceName):
                self._package.remove_resource(resourceName)
            resource.name = resourceName
            #try to get schema
            if isinstance(schema, Schema):
                self._report.addReportDebug(resourceName,"use provided schema") 
                resource.schema = schema
            elif isinstance(schema, str):
                schemaFilename = os.path.join(self._schemaPath,schema)
                if os.access(schemaFilename, os.R_OK):
                    self._report.addReportDebug(resourceName,"validate using schema '{}'".format(schema)) 
                    resource.schema = Schema(schemaFilename)
                else:
                    self._report.addReportError(resourceName,"could not read schema '{}'".format(schemaFilename))
            else:
                self._report.addReportDebug(resourceName,"no recognized schema format")
            #check string types, empty rows and columns
            if (self._adjustTypeForStringColumns and not 
                validate(resource, pick_errors=["type-error"], skip_errors=skip_errors).valid):
                
                stringColumnNames = [field["name"] for field in resource.schema.fields if (field.get("type",None)=="string")]
                self._updateTypeForStringColumns(sheetName, stringColumnNames, headerRows, resourceName)                
            if self._removeEmptyRows and not validate(resource, pick_errors=["blank-row"], skip_errors=skip_errors).valid:
                self._removeEmptyRowsForSheet(sheetName,resourceName)
            if self._removeEmptyColumns and not validate(resource, pick_errors=["extra-label"], skip_errors=skip_errors).valid:
                self._removeEmptyColumnsForSheet(sheetName,resourceName)
            self._report.addReportDebug(resourceName,"validate using schema")    
            if not skip_errors==None:
                self._report.addReportDebug(resourceName,"skip errors in validation: {}".format(str(skip_errors)))
            resource_validation = validate(resource, skip_errors=skip_errors)
            self._report.setReportFrictionless(resourceName, resource_validation)
            self._report.addReportDebug(resourceName,resource_validation.stats)
            if not resource_validation.valid:             
                self._report.addReportError(resourceName,"frictionless validation failed for '{}' sheet".format(sheetName))
            elif skip_errors==None:
                self._report.addReportInfo(resourceName,"succesfull frictionless validation '{}' from '{}' sheet".format(
                    resourceName, sheetName))
            #try to add resource
            self._package.add_resource(resource)   
            #check missing columns for normal run
            if skip_errors==None and not resource_validation.valid:
                self._checkMissingColumns(sheetName,resourceName,headerRows-1)
    
    def _validatePackage(self):
        self._report.addReport("package", False)
        self._report.addReportDebug("package","validate package") 
        package_validation = validate(self._package)
        if not package_validation.valid:
            self._report.addReportError("package","frictionless validation package failed")            
        else:
            self._report.addReportInfo("package","succesfull frictionless validation package")
        self._report.setReportFrictionless("package",package_validation)
        #return package and validation report
        return package_validation
    
    def _validateLogic(self):
        self._report.addReport("logic", False)
        try:
            if not self._allowAdditionalSheets:
                resourceFilename = os.path.join(self._dataPath, self._resourceFilename)
                wb = load_workbook(resourceFilename, data_only = True)
                additionalSheets = set(wb.sheetnames).difference(self._allowedSheets)
                if len(additionalSheets)>0:
                    self._report.addReportError("logic",
                            "unexpected sheet(s) found: '{}'".format(str("', '".join(additionalSheets))))
                else:
                    self._report.addReportDebug("logic","no unexpected sheet(s) found in {}".format(self._resourceFilename))
        except Exception as e:
            self._report.addReportError("logic","problem validating logic: {}".format(str(e)))
            
    def _removeEmptyRowsForSheet(self, sheetName:str, resourceName:str):
        self._report.addReportDebug(resourceName,
                                        "remove empty rows from sheet '{}'".format(sheetName))
        try:
            resourceFilename = os.path.join(self._dataPath, self._resourceFilename)
            if not os.access(resourceFilename, os.W_OK):
                self._report.addReportError(resourceName,
                                            "no write access to {}, can't try to remove empty rows".format(resourceFilename))
            else:
                wb = load_workbook(resourceFilename)
                if not sheetName in wb.sheetnames:
                    self._report.addReportError(resourceName,
                                            "can't find sheet {} in resource".format(sheetName))
                else:
                    ws = wb[sheetName]
                    self._report.addReportDebug(resourceName,
                                            "detect {} columns and {} rows in sheet".format(ws.max_column, ws.max_row))
                    deletablIds = []
                    for row in ws.rows:
                        values = [cell for cell in row if not (cell.internal_value==None or
                                                              str(cell.internal_value).isspace())]
                        if len(values)==0:
                            for cell in row:
                                id = cell.row
                                break
                            deletablIds.append(id)
                    if len(deletablIds)>0:
                        self._report.addReportWarning(resourceName,
                                            "removed {} empty rows from sheet '{}'".format(len(deletablIds),sheetName)) 
                        #try to delete efficient...
                        sortedDeletableIds = sorted(deletablIds, reverse=True)
                        deletableList = []
                        for i in range(len(sortedDeletableIds)):
                            id = sortedDeletableIds[i]
                            deletableList.append(id)
                            if ((i+1)<len(sortedDeletableIds)) and sortedDeletableIds[i+1]==(id-1):
                                #just continue
                                pass
                            else:
                                ws.delete_rows(deletableList[-1],len(deletableList))
                    #always save
                    wb.save(filename = resourceFilename)
                    self._report.addReportDebug(resourceName,
                                            "updated {} after removing {} empty rows".format(
                        self._resourceFilename,len(deletablIds))) 
        except Exception as e:
            self._report.addReportError(resourceName,"problem removing empty rows: {}".format(str(e)))
                    
    def _removeEmptyColumnsForSheet(self, sheetName:str, resourceName:str):
        self._report.addReportDebug(resourceName,
                                        "remove empty columns from sheet '{}'".format(sheetName))
        try:
            resourceFilename = os.path.join(self._dataPath, self._resourceFilename)
            if not os.access(resourceFilename, os.W_OK):
                self._report.addReportError(resourceName,
                    "no write access to {}, can't try to remove empty columns".format(resourceFilename))
            else:
                wb = load_workbook(resourceFilename)
                if not sheetName in wb.sheetnames:
                    self._report.addReportError(resourceName,
                        "can't find sheet {} in resource".format(sheetName))
                else:
                    ws = wb[sheetName]
                    self._report.addReportDebug(resourceName,
                        "detect {} columns and {} rows in sheet".format(ws.max_column, ws.max_row))
                    deletablIds = []
                    for column in ws.columns:
                        values = [cell for cell in column if not (cell.internal_value==None or 
                                                                  str(cell.internal_value).isspace())]
                        if len(values)==0:
                            for cell in column:
                                id = cell.column
                                break
                            deletablIds.append(id)
                    if len(deletablIds)>0:
                        self._report.addReportWarning(resourceName,
                                            "removed {} empty columns from sheet '{}'".format(len(deletablIds),sheetName)) 
                        #try to delete efficient...
                        sortedDeletableIds = sorted(deletablIds, reverse=True)
                        deletableList = []
                        for i in range(len(sortedDeletableIds)):
                            id = sortedDeletableIds[i]
                            deletableList.append(id)
                            if ((i+1)<len(sortedDeletableIds)) and sortedDeletableIds[i+1]==(id-1):
                                #just continue
                                pass
                            else:
                                ws.delete_cols(deletableList[-1],len(deletableList))                
                    #always save
                    wb.save(filename = resourceFilename)
                    self._report.addReportDebug(resourceName,
                                            "updated {} after removing {} empty columns".format(
                        self._resourceFilename,len(deletablIds))) 
        except Exception as e:
            self._report.addReportError(resourceName,"problem removing empty columns: {}".format(str(e)))
            
    def _updateTypeForStringColumns(self, sheetName:str, columnNames: list, headerRows: int, resourceName:str):
        if len(columnNames)>0:
            self._report.addReportDebug(resourceName,
                                            "convert column(s) '{}' to string in '{}'".format(
                                                "', '".join(columnNames),sheetName))
            try:
                resourceFilename = os.path.join(self._dataPath, self._resourceFilename)
                if not os.access(resourceFilename, os.W_OK):
                    self._report.addReportError(resourceName,
                        "no write access to {}, can't try to convert columns to string".format(resourceFilename))
                else:
                    wb = load_workbook(resourceFilename)
                    if not sheetName in wb.sheetnames:
                        self._report.addReportError(resourceName,
                            "can't find sheet {} in resource".format(sheetName))
                    else:
                        ws = wb[sheetName]
                        totalUpdatedNumber=0
                        for column in ws.columns:
                            if column[headerRows-1].value in columnNames:
                                updatedNumber = 0
                                for item in column[headerRows:]:
                                    if not (item.internal_value==None or isinstance(item.value,str)):
                                        item.value = str(item.value)
                                        updatedNumber+=1
                                if updatedNumber>0:
                                    self._report.addReportWarning(resourceName,
                                        "changed type for {} entries from column '{}' in '{}'".format(
                                            updatedNumber,column[0].value,resourceName))
                                totalUpdatedNumber+=updatedNumber
                        #always save
                        wb.save(filename = resourceFilename)
                        self._report.addReportDebug(resourceName,
                                                "updated {} cells to string in {}".format(
                                                totalUpdatedNumber,self._resourceFilename)) 
            except Exception as e:
                self._report.addReportError(resourceName,"problem converting sheet columns to string: {}".format(str(e)))
                
    def _getSheetColumnNames(self, sheetName:str, skipRows=0):
        self._logger.debug("get column names from sheet '{}'".format(sheetName))
        resourceFilename = os.path.join(self._dataPath, self._resourceFilename)
        wb = load_workbook(resourceFilename, data_only=True)
        columnNames = []
        if not sheetName in wb.sheetnames:
            self._logger.error("can't find sheet {} in resource".format(sheetName))
        else:
            ws = wb[sheetName]
            rowCounter = 0
            for row in ws.rows:
                if rowCounter==skipRows:
                    columnNames = [str(cell.internal_value) for cell in row]
                    break
                else:
                    rowCounter+=1
        return columnNames
    
    def createPackageJSON(self, filename:str = None):
        """
        Create frictionless package data
        """
        if not filename==None:
            try:
                return self._package.to_json(filename)
            except:
                self._logger.error("can't store package to '{}'".format(filename))
                return self._package.to_json()
        else:
            return self._package.to_json()
    
    def createReport(self):
        """
        Create JSON object with report
        """
        def createFrictionlessData(fdata):
            list = []
            for task in fdata["tasks"]:
                for error in task["errors"]:
                    list.append({"resource": task["resource"]["name"],
                                 "rowPosition": error.get("rowPosition",None), 
                                 "fieldPosition": error.get("fieldPosition",None),
                                 "code": error.get("code",None),
                                 "description": error.get("description",None)})
            return pd.DataFrame(list)
        def excelCoordinates(row, col):
            try:
                LETTERS = "ABCDEFGHIJKLMNOPQRSTUVWXYZ"
                if not col==None and not np.isnan(col):
                    result = []
                    while col:
                        col, rem = divmod(col-1, 26)
                        result[:0] = LETTERS[int(rem)]
                    if not row==None and not np.isnan(row):
                        return "".join(result) + str(int(row))
                    else:
                        return "column " + "".join(result)
                elif not row ==None and not np.isnan(row):
                    return "row "+str(int(row))
                else:
                    return None
            except:
                return None
        reportObject = {"name": os.path.basename(self._name), "valid": self.valid, "reports": []}
        previousFrictionlessCodes = set()
        for report in self._report.reports:
            isNotValid=((not self._report[report]["valid"]) or 
                        ("frictionless" in self._report[report].keys() and 
                         (not self._report[report]["frictionless"]["valid"])))
            reportDetails = {"name": report, 
                             "valid": False if isNotValid else True,
                             "warnings": [], "errors": [], "frictionless": []}  
            if isNotValid or len(self._report[report]["warnings"])>0:
                if len(self._report[report]["warnings"])>0:
                    for warning in self._report[report]["warnings"]:
                        reportDetails["warnings"].append(warning)
                if len(self._report[report]["errors"])>0:
                    for error in self._report[report]["errors"]:
                        reportDetails["errors"].append(error)
                if "frictionless" in self._report[report].keys():
                    frictionlessData = createFrictionlessData(self._report[report]["frictionless"])
                    if len(frictionlessData)>0:
                        frictionlessCodes = frictionlessData.code.value_counts()
                        if report=="package":
                            #don't repeat frictionless entries from separate sheets
                            frictionlessCodeList = set(frictionlessCodes.index).difference(previousFrictionlessCodes)
                            if len(frictionlessCodeList)==0:
                                continue
                        else:
                            frictionlessCodeList = frictionlessCodes.index
                            previousFrictionlessCodes.update(frictionlessCodes.index)
                        for code in frictionlessCodeList:
                            frictionlessDetails = {"name": code, "details": []}
                            for id,row in frictionlessData[frictionlessData.code==code].iterrows():
                                location = excelCoordinates(row["rowPosition"],row["fieldPosition"])
                                frictionlessDetails["details"].append({"resource": row["resource"], 
                                                                       "location": location,
                                                                       "row": row["rowPosition"],
                                                                       "column": row["fieldPosition"],
                                                                       "description": row["description"]})   
                            reportDetails["frictionless"].append(frictionlessDetails)
            reportObject["reports"].append(reportDetails)
        return reportObject
        
    def createTextReport(self, textWidth=100, examples=3):
        """
        Create text version report
        """
        #create report
        reportObject = self.createReport()
        #create text    
        reportText = "=== {} '{}' ===".format(("VALID" if reportObject["valid"] else "INVALID"),reportObject["name"])
        previousFrictionlessCodes = set()
        for reportDetails in reportObject["reports"]:
            if (not reportDetails["valid"]) or len(reportDetails["warnings"])>0:
                reportText = reportText + "\n** problem with '{}' **".format(reportDetails["name"])
                if len(reportDetails["warnings"])>0:
                    reportText = reportText + "\n- warnings:"
                    for warning in reportDetails["warnings"]:
                        reportText = reportText + "\n" + textwrap.fill(warning,textWidth-4,
                                                    initial_indent="  - ",subsequent_indent="    ")
                if len(reportDetails["errors"])>0:
                    reportText = reportText + "\n- errors:"
                    for error in reportDetails["errors"]:
                        reportText = reportText + "\n"+textwrap.fill(error,textWidth-4,
                                                    initial_indent="  - ",subsequent_indent="    ")
                if len(reportDetails["frictionless"])>0:
                    reportText = reportText + "\n- frictionless report:"
                    for frictionlessItem in reportDetails["frictionless"]:
                        reportText = reportText + "\n  - error type '{}': {}x".format(frictionlessItem["name"],
                                                                                      len(frictionlessItem["details"]))
                        rowCounter = 0
                        for row in frictionlessItem["details"]:
                            if rowCounter>=examples:
                                reportText = reportText + "\n    - ..."
                                break
                            reportText = reportText + "\n" + textwrap.fill("{}{}: {}".format(row["resource"],
                                     ((" ["+row["location"]+"]") if not row["location"]==None else ""),row["description"])
                                     ,textWidth-6,initial_indent="    - ",subsequent_indent="      ")
                            rowCounter+=1
        return reportText
        
    @property
    def valid(self):        
        return self._report.valid
    
    @property
    def reports(self):
        return self._report.reports
    
    def __getitem__(self, key):
        return self._report[key]
        
        
class ValidatePedigree(ValidateGeneral):

    default_checks = [
                {"schema": "pedigree_varieties.schema.json", "sheet": "varieties"},
                {"schema": "pedigree_synonyms.schema.json", "sheet": "synonyms"},
                {"schema": "pedigree_countries.schema.json", "sheet": "countries"},
                {"schema": "pedigree_ancestors.schema.json", "sheet": "ancestors"},
                {"schema": "pedigree_breeders.schema.json", "sheet": "breeders"},
            ]
                 
    def __init__(self, dataPath:str, resourceFilename:str, **args):
        
        super(ValidatePedigree, self).__init__(dataPath, resourceFilename, **args)
        
        #logging
        self._logger = logging.getLogger(__name__)
        
        #initialise
        self._report = ValidateReport(self._logger)
        self._package = Package()
        
        self._name = args.get("name", resourceFilename)
        
        self._availableSheets = set()
        self._allowedSheets = set()

        #first check access to file   
        self._report.addReport("general", False)
        if not os.access(os.path.join(self._dataPath,self._resourceFilename), os.R_OK):
            self._report.addReportError("general","no read access to '{}'".format(self._resourceFilename))
            self._resourceFilename = None       
        else:
            #start processing xlsx
            self._report.addReportDebug("general", "read access to '{}'".format(self._resourceFilename))               
            try:
                try:
                    wb = load_workbook(os.path.join(self._dataPath,self._resourceFilename), data_only = True)
                    self._availableSheets = set(wb.sheetnames)
                except Exception as e:
                    self._report.addReportError("general", 
                                                "problem retrieving sheetnames from '{}': {}".format(
                                                    self._resourceFilename,str(e)))
            except Exception as e:
                self._report.addReportError("general", "problem with output file: {}".format(str(e)))
                
        #now validate
        if self._report.valid:
            
            #validate sheets and create resources
            for item in self.default_checks:
                #check existence and readability schema
                assert os.access(os.path.join(self._schemaPath,item["schema"]), os.R_OK)
                #validate
                self._validateResource(item["sheet"], item["schema"], "pedigree_"+item["sheet"], 1)

            #validate package
            self._validatePackage()

            #validate logic
            self._validateLogic()
        
class ValidateData(ValidateGeneral):

    default_checks = [
                {"schema": "data_default_metadata.schema.json", "sheet": "metadata", 
                 "optional": False, "condition": []},
                {"schema": "data_default_varieties.schema.json", "sheet": "varieties", 
                 "optional": False, "condition": []},
                {"schema": "data_default_experiments.schema.json", "sheet": "experiments", 
                 "optional": True, "condition": []},
                {"schema": "data_default_markers.schema.json", "sheet": "markers", 
                 "optional": True, "condition": ["experiments"]},
                {"schema": "data_default_mappings.schema.json", "sheet": "mappings", 
                 "optional": False, "condition": ["markers"]},
                {"schema": "data_default_scores.schema.json", "sheet": "scores", 
                 "optional": False, "condition": ["markers"]},
                {"schema": "data_default_sequences.schema.json", "sheet": "sequences", 
                 "optional": True, "condition": ["experiments"]},
            ]
            
    def __init__(self, dataPath:str, resourceFilename:str, pedigreePackage:str, **args):
        
        super(ValidateData, self).__init__(dataPath, resourceFilename, **args)
        
        #get config
        self._pedigreePackageName = pedigreePackage
        
        #first check access to file   
        self._report.addReport("general", False)
        if not os.access(os.path.join(self._dataPath,self._resourceFilename), os.R_OK):
            self._report.addReportError("general","no read access to '{}'".format(self._resourceFilename))
            self._resourceFilename = None       
        else:
            #first include pedigree
            self._pedigreePackage = Package(os.path.join(self._dataPath,self._pedigreePackageName))
            self._report.addReport("pedigree", False)
            self._report.addReportDebug("pedigree", "read access to '{}'".format(self._pedigreePackage))               
            try:
                if self._pedigreePackage.has_resource("pedigree_varieties"):
                    self._package.add_resource(self._pedigreePackage.get_resource("pedigree_varieties"))
                else:
                    self._report.addReportError("pedigree", "no varieties in pedigree")
                if self._pedigreePackage.has_resource("pedigree_countries"):
                    self._package.add_resource(self._pedigreePackage.get_resource("pedigree_countries"))
                else:
                    self._report.addReportError("pedigree", "no countries in pedigree")
                if self._pedigreePackage.has_resource("pedigree_breeders"):
                    self._package.add_resource(self._pedigreePackage.get_resource("pedigree_breeders"))
                else:
                    self._report.addReportError("pedigree", "no breeders in pedigree")
            except Exception as e:
                self._report.addReportError("pedigree", "problem with pedigree: {}".format(str(e)))
            #now start processing xlsx
            self._report.addReportDebug("general", "read access to '{}'".format(self._resourceFilename))               
            try:
                try:
                    wb = load_workbook(os.path.join(self._dataPath,self._resourceFilename), data_only = True)
                    self._availableSheets = set(wb.sheetnames)
                except Exception as e:
                    self._report.addReportError("general", 
                            "problem retrieving sheetnames from '{}': {}".format(self._resourceFilename,str(e)))
            except Exception as e:
                self._report.addReportError("general", "problem with output file: {}".format(str(e)))
                
        #now validate
        if self._report.valid:
            
            #validate sheets and create resources
            for item in self.default_checks:
                #not everything is (always) required or expected...
                validateThisSheet = False
                createEmptySheet = False
                if not item["optional"]:
                    if len(item["condition"])==0:
                        validateThisSheet = True
                    elif len(self._availableSheets.intersection(item["condition"]))>0:
                        validateThisSheet = True
                    else:
                        validateThisSheet = False
                elif item["sheet"] in self._availableSheets:
                    if len(item["condition"])==0:
                        validateThisSheet = True
                    elif len(self._availableSheets.intersection(item["condition"]))>0:
                        validateThisSheet = True
                    else:
                        validateThisSheet = False
                else:
                    validateThisSheet = False
                    if len(item["condition"])>0:
                        if len(self._availableSheets.intersection(item["condition"]))>0:
                            createEmptySheet = True
                #validate                        
                if validateThisSheet:
                    #check existence and readability schema
                    assert os.access(os.path.join(self._schemaPath,item["schema"]), os.R_OK)
                    #validate
                    self._validateResource(item["sheet"], item["schema"], item.get("name", item["sheet"]),
                                           item.get("headerRows",1))
                elif createEmptySheet:
                    #check existence and readability schema
                    assert os.access(os.path.join(self._schemaPath,item["schema"]), os.R_OK)
                    self._createEmptyResource(item["sheet"], item["schema"], item.get("name", item["sheet"]))
                    
                    
            #check marker scores
            if (self._package.has_resource("scores") and self._package.has_resource("markers") 
                and self._package.has_resource("varieties")):
                scores = pd.DataFrame(extract(self._package.get_resource("scores")))
                markers = pd.DataFrame(extract(self._package.get_resource("markers")))
                if len(markers)>0:
                    markers = markers.set_index("id")
                varieties = pd.DataFrame(extract(self._package.get_resource("varieties")))
                if len(varieties)>0:
                    varieties = varieties.set_index("id")        
                for score_id,score_row in scores.iterrows():
                    scoreFilename = score_row["source"]
                    self._report.addReport(score_row["source"], False)
                    fullScoreFilename = os.path.join(self._dataPath, scoreFilename)
                    if not os.access(fullScoreFilename, os.R_OK):
                        self._report.addReportError(score_row["source"], 
                                        "no read access to '{}'".format(fullScoreFilename))
                    else:
                        open_fn = gzip.open if fullScoreFilename.endswith(".gz") else open
                        with open_fn(fullScoreFilename, "rt") as f: 
                            reader = csv.reader(f, delimiter=",")
                            line = list(next(reader))
                            markerSet = set()
                            markerStringIndex = [str(item) for item in markers.index]
                            varietyStringIndex = [str(item) for item in varieties.index]
                            for item in list(map(str,line[1:])):
                                if item in markerStringIndex:
                                    if not item in markerSet:
                                        markerSet.add(item)
                                    else:
                                        self._report.addReportError(score_row["source"], 
                                                "marker '{}' used multiple times".format(item))
                                else:
                                    self._report.addReportError(score_row["source"], 
                                                "unexpected marker '{}'".format(item))
                            for line in reader:
                                if not str(line[0]) in varietyStringIndex:
                                    self._report.addReportError(score_row["source"], 
                                                "unexpected variety '{}'".format(line[0]))

            #validate package
            self._validatePackage()

            #validate logic
            self._validateLogic()
